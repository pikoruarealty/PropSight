"""Generate synthetic sample workbooks for manual QA. Not real client data.

Run:  python sample_data/generate_samples.py
Deterministic (seeded) so E2E checks can be cross-verified independently.
"""

from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

OUT_DIR = Path(__file__).parent
rng = random.Random(42)

HEADERS = [
    "Name", "Email address", "Duplicacy Check", "Phone number", "Stage",
    "lead Source", "Form", "First Call Date", "Latest Call Date",
    "Latest Call Status", "Visit status", "Buying Status", "Interest Level",
    "Budget", "Purpose of Buying", "Configuration Required",
    "Location Preference", "Qualitative Remarks", "Follow Up Date",
    "Business Profile", "HWC", "Properties Seen",
    "Current Location (if from Ahmedabad)", "Marital Status", "Kids",
    "Family Size", "Looking Since", "Current Size", "Reason for Shifting",
    "Current City", "Form", "Lead for Project",
]

FIRST = ["Amit", "Priya", "Rahul", "Sneha", "Vikram", "Neha", "Karan", "Pooja",
         "Sanjay", "Ritu", "Manish", "Kavita", "Deepak", "Anjali", "Rohan"]
LAST = ["Patel", "Shah", "Mehta", "Desai", "Joshi", "Trivedi", "Vyas", "Parikh"]
STAGES = ["New", "Contacted", "Visit Scheduled", "Visit Done", "Negotiation", "Closed Won", "Closed Lost"]
SOURCES = ["Google Ads", "Facebook", "99acres", "MagicBricks", "Walk-in", "Referral"]
FORMS = ["Website Enquiry", "Landing Page A", "Landing Page B", "Call-in"]
FORMS_2 = ["FB-Lead-Gen-Mar", "GG-Search-Brand", "Organic", None, None]
CALL_STATUS = ["Connected", "Not Picking", "Switched Off", "Call Back Later"]
VISIT = ["Visit Done", "Scheduled", "Not Visited", "Cancelled"]
BUYING = ["Ready to Buy", "Thinking", "Not Interested", "Booked", "Follow Up"]
INTEREST = ["Hot", "Warm", "Cold", "Medium"]
PURPOSE = ["Self Use", "Investment", "Both", None]
LOCATIONS = ["Bopal", "South Bopal", "Satellite", "Thaltej", "Shela", "Gota", "Vastrapur"]
# Deliberate near-duplicate spellings to exercise fuzzy dedupe.
CITIES = ["Ahmedabad", "Ahmedabad", "Ahmedabad", "Ahmadabad", "Amdavad", "Mumbai", "Surat", "Rajkot", None]
MARITAL = ["Married", "Single", "married", None]
PROFILES = ["Textile business owner", "Software Engineer at TCS", "Doctor - own clinic",
            "Chartered Accountant", "Govt employee (GPSC)", "NRI - settled in USA",
            "Retired bank officer", "Share market trader", "Runs a pharma firm", None]
REASONS = ["Growing family, need more space", "Kids school is far, want better school district",
           "Investment for rental income", "Job transfer to Ahmedabad",
           "Want bigger house with garden", "Current flat too small after second baby",
           "Upgrading lifestyle", None, None]
LOOKING = ["1 month", "3 months", "6 months", "1 year", "2 years", "15 days", None]
REMARKS = ["Very keen, call after Diwali", "Wants corner unit only", "Price sensitive",
           "Comparing with 2 other projects", None, None, None]
# Real HWC values are freeform notes, not a fixed enum — any non-empty value marks the client.
HWC_VALUES = ["Yes", "High priority", "VIP", "Referred by owner", None, None, None, None]

PROJECT_MAP = {
    "Apartment": ["Skyline Heights", "Green Residency"],
    "Bungalow": ["Royal Villas", "Orchid Greens"],
    "Plot": ["Sunrise Plots"],
}
SEEN_MAP = {
    "Apartment": ["Godrej Garden City", "Adani Shantigram", "Skyline Heights"],
    "Bungalow": ["Aaryan Villas", "Godrej Garden City"],
    "Plot": ["Arvind Uplands", "Adani Shantigram"],
}
BUDGETS = {
    "Apartment": ["45 L", "60 Lakh", "75 L", "80-90 lakhs", "1 Cr", "55 L", None],
    "Bungalow": ["1.5 Cr", "2 Cr", "1 Cr - 2 Cr", "2.5cr", "3 Cr", None],
    "Plot": ["30 L", "40 L", "25 Lakh", "50 L", "60-70 lakhs", None],
}
CONFIGS = {
    "Apartment": ["2 BHK", "3 BHK", "3 BHK", "4 BHK", None],
    "Bungalow": ["4 BHK", "5 BHK", "4 BHK", None],
    "Plot": [None, None, "3 BHK"],
}


def _date(base: date, spread: int, fmt: str) -> str | date:
    d = base + timedelta(days=rng.randint(0, spread))
    if fmt == "iso":
        return d.isoformat()
    if fmt == "dmy":
        return d.strftime("%d/%m/%Y")
    return d  # real date cell


def make_row(ptype: str, sparse: bool) -> list:
    """One synthetic lead. `sparse` rows keep only a handful of fields."""
    name = f"{rng.choice(FIRST)} {rng.choice(LAST)}"
    if sparse:
        row = {h: None for h in range(len(HEADERS))}
        values = [None] * len(HEADERS)
        values[0] = name
        values[3] = f"98{rng.randint(10000000, 99999999)}"
        values[5] = rng.choice(SOURCES)
        return values
    fmt = rng.choice(["iso", "dmy", "cell"])
    first_call = date(2024, rng.randint(1, 6), rng.randint(1, 28))
    family = rng.choice([2, 3, 4, 5, 6, None])
    kids = rng.choice(["0", "1", "2", "3", None]) if family else None
    values = [
        name,
        f"{name.split()[0].lower()}.{rng.randint(1,99)}@example.com" if rng.random() > 0.3 else None,
        "Duplicate" if rng.random() < 0.12 else rng.choice(["No", None]),
        f"98{rng.randint(10000000, 99999999)}",
        rng.choice(STAGES),
        rng.choice(SOURCES),
        rng.choice(FORMS),
        _date(first_call, 0, fmt),
        _date(first_call, 45, fmt) if rng.random() > 0.25 else None,
        rng.choice(CALL_STATUS),
        rng.choice(VISIT),
        rng.choice(BUYING),
        rng.choice(INTEREST),
        rng.choice(BUDGETS[ptype]),
        rng.choice(PURPOSE),
        rng.choice(CONFIGS[ptype]),
        rng.choice(LOCATIONS),
        rng.choice(REMARKS),
        _date(date(2024, 7, 1), 120, "iso") if rng.random() > 0.5 else None,
        rng.choice(PROFILES),
        rng.choice(HWC_VALUES),
        ", ".join(rng.sample(SEEN_MAP[ptype], rng.randint(1, 2))) if rng.random() > 0.5 else None,
        rng.choice(LOCATIONS) if rng.random() > 0.4 else None,
        rng.choice(MARITAL),
        kids,
        family,
        rng.choice(LOOKING),
        rng.choice(["1 BHK", "2 BHK", "2 BHK", "3 BHK", None]),
        rng.choice(REASONS),
        rng.choice(CITIES),
        rng.choice(FORMS_2),
        rng.choice(PROJECT_MAP[ptype] + [None]),
    ]
    return values


# A sheet with no lead-shaped headers at all — e.g. a monthly target tracker
# some team keeps in the same workbook. Must default to excluded on upload.
JUNK_HEADERS = ["Month", "Target", "Achieved", "Team", "Notes"]
JUNK_TEAMS = ["Team A", "Team B", "Team C"]


def make_junk_row(month: str) -> list:
    return [month, rng.randint(10, 40), rng.randint(5, 35), rng.choice(JUNK_TEAMS), None]


def build(filename: str, ptype: str, sheets: dict[str, int], junk_sheet: str | None = None) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, n_rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(HEADERS)
        for i in range(n_rows):
            ws.append(make_row(ptype, sparse=(i % 7 == 6)))  # every 7th row is sparse
    if junk_sheet:
        ws = wb.create_sheet(title=junk_sheet)
        ws.append(JUNK_HEADERS)
        for month in ["Jan", "Feb", "Mar", "Apr"]:
            ws.append(make_junk_row(month))
    wb.save(OUT_DIR / filename)
    total_sheets = len(sheets) + (1 if junk_sheet else 0)
    print(f"wrote {filename}: {sum(sheets.values())} lead rows in {total_sheets} sheet(s)")


if __name__ == "__main__":
    build("Apartment Leads.xlsx", "Apartment", {"Jan-Mar 2024": 18, "Apr-Jun 2024": 12}, junk_sheet="Sales Targets")
    build("Bungalow Leads.xlsx", "Bungalow", {"Enquiries": 22})
    build("Plot Leads.xlsx", "Plot", {"Q1": 10, "Q2": 8})
    # Undetectable filename — must force manual property-type selection in the UI.
    build("Q1 Export.xlsx", "Apartment", {"Data": 6})
